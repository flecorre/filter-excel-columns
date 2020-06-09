#!/usr/bin/python3

import os
import openpyxl
import logging
import sys
import time

from kivy.app import App
from kivy.uix.gridlayout import GridLayout
from kivy.uix.button import Button
from kivy.core.window import Window
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.checkbox import CheckBox
from kivy.uix.popup import Popup

# LOGGER
logging.basicConfig(
    format='%(asctime)s %(levelname)-8s %(message)s',
    level=logging.INFO,
    datefmt='%Y-%m-%d %H:%M:%S')


class ExcelFilter:

    def __init__(self):
        self.first_range = []
        self.second_range = []
        self.percentage_threshold = None
        self.skip_background = False
        self.skip_normalization = False
        self.excel_output_file = None
        self.report_file_wrong = None
        self.report_file_good = None
        self.workbook = None

    def process_excel_file(self, file, threshold, first_range, second_range, skip_background, skip_normalization):
        self.first_range = first_range
        self.second_range = second_range
        self.percentage_threshold = threshold
        self.skip_background = skip_background
        self.skip_normalization = skip_normalization
        if self.verify_file is True:
            with open(file) as fp:
                line = fp.readline().rstrip('\n')
                while line:
                    self.main(line)
                    line = fp.readline().rstrip('\n')
        else:
            self.main(file)

    @staticmethod
    def verify_file(file):
        if file.lower().endswith(Constant.TXT):
            return True
        else:
            return False

    def main(self, excel_file):
        self.prepare_output_files(excel_file)
        self.open_excel_file(excel_file)
        if not self.skip_background:
            self.subtract_background(self.workbook)
        self.filter_columns(self.workbook)
        if not self.skip_normalization:
            self.calculate_mean_and_normalize_roi(self.workbook, Constant.SHEET_GOOD_ROI, Constant.MEAN_GOOD_ROI)
            self.calculate_mean_and_normalize_roi(self.workbook, Constant.SHEET_WRONG_ROI, Constant.MEAN_WRONG_ROI)
        logging.info("writing processed data to: '{}'".format(self.excel_output_file))
        self.workbook.save(self.excel_output_file)

    def prepare_output_files(self, file):
        self.check_file_extension(file)
        filtered_threshold = Constant.THRESHOLD + str(self.percentage_threshold)
        self.excel_output_file = self.create_output_excel_file_name(file, filtered_threshold)

    @staticmethod
    def check_file_extension(excel_file):
        if not excel_file.lower().endswith((Constant.XLS, Constant.XLSX)):
            raise TypeError('Not an excel file: {}'.format(excel_file))

    @staticmethod
    def create_output_excel_file_name(file, suffix):
        filename = file.split(".")[0]
        file_extension = file.split(".")[1]
        timestamp = time.strftime("%Y-%m-%d-%H-%M-%S")
        name = '{}_{}_{}.{}'.format(filename, suffix, timestamp, file_extension)
        return name

    def open_excel_file(self, excel_file):
        logging.info("opening '{}'...".format(excel_file))
        self.workbook = openpyxl.load_workbook(excel_file)

    def subtract_background(self, wb):
        sheet = wb.active
        subtracted_bg_sheet = self.copy_worksheet(wb, sheet, Constant.SHEET_BACKGROUND_SUBTRACTED)
        logging.info("subtracting background...")
        for row in subtracted_bg_sheet.iter_rows(min_row=Constant.BACKGROUND_MIN_ROW,
                                                 min_col=Constant.BACKGROUND_COLUMN_INDEX):
            background_cell = row[0].value
            for cell in row:
                if cell.value is not None:
                    result = cell.value - background_cell
                    if result != 0:
                        cell.value = result
                    elif result == 0 and cell.column != Constant.BACKGROUND_COLUMN_INDEX:
                        msg_duplicate_bg_row = "/!\ ERROR: COLUMN {} IS THE SAME AS BACKGROUND COLUMN".format(
                            cell.column)
                        sys.exit(msg_duplicate_bg_row)
        subtracted_bg_sheet.delete_cols(Constant.BACKGROUND_COLUMN_INDEX)

    @staticmethod
    def copy_worksheet(wb, sheet, title):
        new_worksheet = wb.copy_worksheet(sheet)
        new_worksheet.title = title
        return new_worksheet

    def filter_columns(self, wb):
        if not self.skip_background:
            sheet = wb[Constant.SHEET_BACKGROUND_SUBTRACTED]
        else:
            sheet = wb.active
        FILTER_MAX_COL = sheet.max_column

        # ITERATE THROUGH COLUMNS AND IDENTIFY BAD COLUMNS GIVEN THE THRESHOLD PERCENTAGE
        logging.info("filtering ROIs...")
        columns_index_wrong = []
        columns_index_good = []
        columns_info_wrong = {}
        columns_info_good = {}
        for col in sheet.iter_cols(min_row=Constant.FILTER_MIN_ROW, min_col=Constant.FILTER_MIN_COL,
                                   max_row=Constant.FILTER_MAX_ROW, max_col=FILTER_MAX_COL):
            if col[0].value is not None:
                first_mean = self.get_mean_from_range_of_rows(col, self.first_range)
                second_mean = self.get_mean_from_range_of_rows(col, self.second_range)
                difference = self.calculate_percentage_difference(first_mean, second_mean)
                if difference > self.percentage_threshold or difference < 0:
                    columns_index_wrong.append(col[0].column)
                    columns_info_wrong.update({col[0].value: difference})
                else:
                    columns_index_good.append(col[0].column)
                    columns_info_good.update({col[0].value: difference})

        # DELETE COLUMNS IF WRONG COLUMNS ARE FOUND
        if len(columns_index_wrong) != 0:
            logging.info("{} good columns found...".format(str(len(columns_index_good))))
            logging.info("{} wrong columns found...".format(str(len(columns_index_wrong))))
            logging.info("deleting columns...")
            sheet_good_roi = self.copy_worksheet(wb, sheet, Constant.SHEET_GOOD_ROI)
            sheet_wrong_roi = self.copy_worksheet(wb, sheet, Constant.SHEET_WRONG_ROI)
            self.delete_columns(sheet_good_roi, columns_index_wrong)
            self.delete_columns(sheet_wrong_roi, columns_index_good)
            # CREATE NEW SHEETS TO WRITE PERCENTAGE CALCULATION RESULTS
            result_above_threshold_title = "{} {} %".format(Constant.RESULT_ABOVE, str(self.percentage_threshold))
            result_below_threshold_title = "{} {} %".format(Constant.RESULT_BELOW, str(self.percentage_threshold))
            wb.create_sheet(result_above_threshold_title)
            wb.create_sheet(result_below_threshold_title)
            self.write_data(wb[result_above_threshold_title], columns_info_wrong)
            self.write_data(wb[result_below_threshold_title], columns_info_good)
        else:
            logging.critical("no column to delete...")

    @staticmethod
    def get_mean_from_range_of_rows(column, list_range):
        sum_rows_from_range = 0
        number_of_rows = 0
        for x in range(list_range[0], list_range[1]):
            sum_rows_from_range += column[Constant.FILTER_MIN_ROW + x - 1].value
            number_of_rows += 1
        return sum_rows_from_range / number_of_rows

    @staticmethod
    def calculate_percentage_difference(first_value, second_value):
        try:
            result = (abs(first_value - second_value) / second_value) * 100.0
        except ZeroDivisionError:
            result = 0
        return result

    @staticmethod
    def delete_columns(sheet, columns_list):
        index = 0
        for column in columns_list:
            sheet.delete_cols(column - index)
            index = index + 1

    @staticmethod
    def write_data(sheet, columns_dict):
        next_row = 1
        for k, v in columns_dict.items():
            sheet.cell(column=1, row=next_row, value=k)
            sheet.cell(column=2, row=next_row, value=v)
            next_row += 1

    def calculate_mean_and_normalize_roi(self, wb, sheet_to_calculate, title_for_new_mean_sheet):
        normalized_sheet_title = "{} normalized".format(sheet_to_calculate)
        selected_sheet = self.copy_worksheet(wb, wb[sheet_to_calculate], normalized_sheet_title)
        min_col = 2
        max_col = selected_sheet.max_column
        min_row = 0
        min_row_mean_calculation = 22
        max_row_mean_calculation = 41
        max_row_normalization = selected_sheet.max_row
        columns_mean = {}
        logging.info("calculating means and normalizing {}...".format(sheet_to_calculate))
        logging.info("mean minimum row: {}...".format(min_row_mean_calculation))
        logging.info("mean maximum row: {}...".format(max_row_mean_calculation))
        # ITERATE A FIRST TIME TO CALCULATE THE MEAN
        for col in selected_sheet.iter_cols(min_row=min_row, min_col=min_col, max_row=max_row_mean_calculation,
                                            max_col=max_col):
            if col[0].value is not None:
                sum_roi_value = 0
                number_roi_values = 0
                for cell in col:
                    # CONDITION NEEDED TO SKIP THE COLUMN TITLE
                    if cell.row >= min_row_mean_calculation:
                        sum_roi_value += cell.value
                        number_roi_values += 1
                mean = (sum_roi_value / number_roi_values)
                columns_mean.update({col[0].value: mean})
        # Iterate a second time to normalize
        for col in selected_sheet.iter_cols(min_row=min_row, min_col=min_col, max_row=max_row_normalization,
                                            max_col=max_col):
            if col[0].value is not None:
                for cell in col:
                    column_title = col[0].value
                    if not cell.value == column_title:
                        cell.value = self.normalize_selected_value(cell.value, column_title, columns_mean)
        wb.create_sheet(title_for_new_mean_sheet)
        self.write_data(wb[title_for_new_mean_sheet], columns_mean)

    @staticmethod
    def copy_worksheet(wb, sheet, title):
        new_worksheet = wb.copy_worksheet(sheet)
        new_worksheet.title = title
        return new_worksheet

    @staticmethod
    def normalize_selected_value(value, target_roi_column, dict_of_means):
        for k in dict_of_means:
            if k == target_roi_column:
                mean = dict_of_means.get(k)
                return (value - mean) / mean


class Constant:
    STARS = "******************************************************"
    XLS = ".xls"
    XLSX = ".xlsx"
    THRESHOLD = "threshold-"
    TXT = ".txt"
    SHEET_BACKGROUND_SUBTRACTED = "bg_subtracted"
    SHEET_GOOD_ROI = "good ROI"
    SHEET_WRONG_ROI = "wrong ROI"
    MEAN_GOOD_ROI = "mean good ROI"
    MEAN_WRONG_ROI = "mean wrong ROI"
    RESULT_ABOVE = "result above"
    RESULT_BELOW = "result below"
    BACKGROUND_MIN_ROW = 2
    BACKGROUND_COLUMN_INDEX = 2
    FILTER_MIN_ROW = 0
    FILTER_MAX_ROW = 41
    FILTER_MIN_COL = 2
    # ~~ FILTER_MAX_COL is set automatically below before filtering


class MyGrid(GridLayout):
    first_row_value = None
    skip_bg_subtraction = False
    skip_normalization = False
    threshold = 0
    first_range_from = 0
    first_range_to = 0
    second_range_from = 0
    second_range_to = 0
    filename = ''
    excel_processor = None

    def __init__(self, **kwargs):
        super(MyGrid, self).__init__(**kwargs)

        self.excel_processor = ExcelFilter()

        self.cols = 1
        self.row_default_height = '300dp'
        self.row_force_default = True
        self.rows = 2

        top_layout = GridLayout(cols=2, row_force_default=True, row_default_height='40dp')
        top_layout.add_widget(Label(text="Threshold", size_hint_x=None, height='20dp', width='200dp'))
        threshold_input = TextInput(text='', font_size=40, height='20dp', multiline=False, write_tab=False)
        threshold_input.bind(text=self.on_threshold)
        top_layout.add_widget(threshold_input)

        top_layout.add_widget(Label(text="First range (row number)", size_hint_x=None, height='200dp', width='200dp'))
        first_range_layout = GridLayout(cols=4, row_force_default=True, row_default_height='40dp')
        first_range_layout.add_widget(Label(text="from", size_hint_x=None, height='20dp'))
        first_range_from = TextInput(font_size=40, height='20dp', multiline=False, write_tab=False)
        first_range_from.bind(text=self.on_first_range_from)
        first_range_layout.add_widget(first_range_from)
        first_range_layout.add_widget(Label(text="to", size_hint_x=None, height='20dp'))
        first_range_to = TextInput(font_size=40, height='20dp', multiline=False, write_tab=False)
        first_range_to.bind(text=self.on_first_range_to)
        first_range_layout.add_widget(first_range_to)
        top_layout.add_widget(first_range_layout)

        top_layout.add_widget(Label(text="Second range (row number)", size_hint_x=None, height='20dp', width='200dp'))
        second_range_layout = GridLayout(cols=4, row_force_default=True, row_default_height='40dp')
        second_range_layout.add_widget(Label(text="from", size_hint_x=None, height='20dp'))
        second_range_from = TextInput(font_size=40, height='20dp', multiline=False, write_tab=False)
        second_range_from.bind(text=self.on_second_range_from)
        second_range_layout.add_widget(second_range_from)
        second_range_layout.add_widget(Label(text="to", size_hint_x=None, height='20dp'))
        second_range_to = TextInput(font_size=40, height='20dp', multiline=False, write_tab=False)
        second_range_to.bind(text=self.on_second_range_to)
        second_range_layout.add_widget(second_range_to)
        top_layout.add_widget(second_range_layout)

        top_layout.add_widget(Label(text="Skip background subtraction", size_hint_x=None, height='20dp', width='200dp'))
        bg_subtraction = CheckBox(active=False)
        bg_subtraction.bind(active=self.bg_subtraction_active)
        top_layout.add_widget(bg_subtraction)

        top_layout.add_widget(Label(text="Skip normalization", size_hint_x=None, height='20dp', width='200dp'))
        bg_subtraction = CheckBox(active=False)
        bg_subtraction.bind(active=self.normalization_active)
        top_layout.add_widget(bg_subtraction)

        top_layout.add_widget(Label(text="File location", size_hint_x=None, height='20dp', width='200dp'))
        self.dragged_file = Label(text='')
        self.dragged_file.font_size = '12dp'
        top_layout.add_widget(self.dragged_file)

        self.add_widget(top_layout)

        bottom_layout = GridLayout(cols=2, row_force_default=True, row_default_height='20dp')
        self.process_btn = Button(text="No file to process",
                                  font_size="20sp",
                                  background_color=(0, 2, 3, 1),
                                  color=(1, 1, 1, 1),
                                  size=(32, 32),
                                  size_hint=(.2, .2),
                                  pos=(300, 250),
                                  disabled=True)
        self.process_btn.bind(on_press=self.on_press)
        bottom_layout.add_widget(self.process_btn)

        self.clear_btn = Button(text="Clear file",
                                font_size="20sp",
                                background_color=(1, .3, .4, .85),
                                color=(1, 1, 1, 1),
                                size=(32, 32),
                                size_hint=(.2, .2),
                                pos=(300, 250),
                                disabled=True)

        self.clear_btn.bind(on_press=self.on_clear)
        bottom_layout.add_widget(self.clear_btn)

        self.add_widget(bottom_layout)

        Window.clearcolor = (.46, .49, .49, 1)
        Window.bind(on_dropfile=self.on_file_drop)

    def on_press(self, instance):
        self.validate_inputs()

    def on_clear(self, instance):
        self.dragged_file.text = ''
        self.process_btn.disabled = True
        self.clear_btn.disabled = True
        self.process_btn.text = 'No file to process'
        self.filename = self.dragged_file.text
        logging.info('file: {}'.format(self.filename))

    def on_threshold(self, instance, value):
        self.threshold = value
        logging.info('threshold: ' + self.threshold)

    def on_first_range_from(self, instance, value):
        self.first_range_from = value
        logging.info('first_range_from: ' + self.first_range_from)

    def on_first_range_to(self, instance, value):
        self.first_range_to = value
        logging.info('first_range_to: ' + self.first_range_to)

    def on_second_range_from(self, instance, value):
        self.second_range_from = value
        logging.info('second_range_from: ' + self.second_range_from)

    def on_second_range_to(self, instance, value):
        self.second_range_to = value
        logging.info('second_range_to: ' + self.second_range_to)

    def on_file_drop(self, window, file_path):
        self.dragged_file.text = r'{}'.format(file_path.decode('utf-8'))
        self.process_btn.disabled = False
        self.clear_btn.disabled = False
        self.process_btn.text = 'Process the file'
        self.filename = self.dragged_file.text
        logging.info('file: {}'.format(self.filename))

    def bg_subtraction_active(self, checkboxInstance, isActive):
        if isActive:
            self.skip_bg_subtraction = True
        else:
            self.skip_bg_subtraction = False
        logging.info('skip bg subtraction: ' + str(self.skip_bg_subtraction))

    def normalization_active(self, checkboxInstance, isActive):
        if isActive:
            self.skip_normalization = True
        else:
            self.skip_normalization = False
        logging.info('skip bg subtraction: ' + str(self.skip_normalization))

    def validate_threshold(self):
        try:
            if not (0 <= int(self.threshold) <= 100):
                raise ValueError
            else:
                logging.error('threshold is valid: {}'.format(self.threshold))
        except ValueError:
            msg = 'Threshold value should be a number between 0 and 100, current value: {}'.format(self.threshold)
            logging.error(msg)
            return msg

    def validate_first_range(self):
        try:
            if int(self.first_range_from) > int(self.first_range_to):
                return 'First range error: {} is greater than {}!'.format(self.first_range_from, self.first_range_to)
            if int(self.first_range_from) == int(self.first_range_to):
                return 'First range error: there is no valid interval!'
            if int(self.first_range_from) == 1:
                return 'First range error: row 1 is column title!'
        except ValueError:
            msg = 'First range not valid numbers'
            logging.error(msg)
            return msg

    def validate_second_range(self):
        try:
            if int(self.second_range_from) > int(self.second_range_to):
                return 'Second range error: {} is greater than {}!'.format(self.second_range_from, self.second_range_to)
            if int(self.second_range_from) == int(self.second_range_to):
                return 'Second range error: there is not interval!'
            if int(self.second_range_from) <= int(self.first_range_to):
                return 'Error: first range is greater than second row!'
        except ValueError:
            msg = 'Second range not valid numbers'
            logging.error(msg)
            return msg

    def validate_excel_file_path(self):
        if not os.path.isfile(self.filename):
            logging.info('Error looking for excel file: {}'.format(self.filename))
            return 'Excel file not found, drag again!'
        else:
            logging.info('Excel file found: {}'.format(self.filename))

    def validate_inputs(self):
        errors = []
        threshold_error = self.validate_threshold()
        first_range_error = self.validate_first_range()
        second_range_error = self.validate_second_range()
        excel_file_error = self.validate_excel_file_path()
        if threshold_error is not None:
            errors.append(threshold_error)
        if first_range_error is not None:
            errors.append(first_range_error)
        if second_range_error is not None:
            errors.append(second_range_error)
        if excel_file_error is not None:
            errors.append(excel_file_error)
        if errors:
            self.display_error_popup(errors)
        else:
            first_range = [int(self.first_range_from), int(self.first_range_to)]
            second_range = [int(self.second_range_from), int(self.second_range_to)]
            try:
                self.excel_processor.process_excel_file(self.filename, int(self.threshold), first_range, second_range,
                                                        self.skip_bg_subtraction, self.skip_normalization)
                self.display_done_popup()
            except Exception as e:
                errors.append(str(e))
                self.display_error_popup(errors)

    @staticmethod
    def display_error_popup(errors):
        layout = GridLayout(cols=1, padding=10)

        for e in errors:
            popup_label = Label(text=e)
            layout.add_widget(popup_label)
        close_button = Button(text="Close the pop-up")
        layout.add_widget(close_button)
        popup = Popup(title='Error',
                      content=layout,
                      size_hint=(None, None), size=('700dp', '300dp'))
        popup.open()
        close_button.bind(on_press=popup.dismiss)

    @staticmethod
    def display_done_popup():
        layout = GridLayout(cols=1, padding=10)

        popup_label = Label(text='DONE!')
        layout.add_widget(popup_label)
        close_button = Button(text="Close the pop-up")
        layout.add_widget(close_button)
        popup = Popup(title='All done',
                      content=layout,
                      size_hint=(None, None), size=('700dp', '200dp'))
        popup.open()
        close_button.bind(on_press=popup.dismiss)


class FilterExcelProgram(App):
    def build(self):
        return MyGrid()


if __name__ == '__main__':
    FilterExcelProgram().run()
